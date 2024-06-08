import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Variables
jobInfoStartingCell = 'G4';
dataCell = 'B10';
multiLevelNumCell = 'A10';
multiLevelNum = '0';

def increment_cell(cell):
    # Separate the column letters and the row number
    column = ''.join([c for c in cell if c.isalpha()])
    row = ''.join([c for c in cell if c.isdigit()])
    
    # Increment the row number by 1
    new_row = str(int(row) + 1)
    
    # Combine the column and new row number
    new_cell = column + new_row
    return new_cell

def incrementMultiLevelNum(num, level):
    if level == 'sub':
        if num.count('.')==0:
            num=num+'.1'
            return num
        else:
            return num.split('.')[0]+'.'+str(int(num.split('.')[1])+1)
    elif level == 'subsub':
        if num.count('.')==0:
            num=num+'.1.1'
            return num
        if num.count('.')==1:
            num=num+'.1'
            return num
        else:
            return num.split('.')[0]+'.'+num.split('.')[1]+'.'+str(int(num.split('.')[2])+1)
    elif level == 'main':
        num = num.split('.')[0]
        return str(int(num) + 1)

# Object Defintion
class Project:
    def __init__(self, client, jobNumber, jobName, LSD, date, equipment):
        self.client = client
        self.jobNumber = jobNumber
        self.jobName = jobName
        self.LSD = LSD
        self.date = date
        self.equipment = equipment 

# Creating an instance of the Project class
project = Project(client="Pembina Gas Services", 
                  jobNumber="21-230", 
                  jobName="Gas Lift", 
                  LSD="11-05", 
                  date="2024-03-10", 
                  equipment="Heat Trace Controllers")

# Accessing the attributes of the project object
cells = {
    jobInfoStartingCell[0] + str(int(jobInfoStartingCell[1:])+0): project.client,
    jobInfoStartingCell[0] + str(int(jobInfoStartingCell[1:])+1): project.jobNumber,
    jobInfoStartingCell[0] + str(int(jobInfoStartingCell[1:])+2): project.jobName,
    jobInfoStartingCell[0] + str(int(jobInfoStartingCell[1:])+3): project.LSD,
    jobInfoStartingCell[0] + str(int(jobInfoStartingCell[1:])+4): project.date,
    jobInfoStartingCell[0] + str(int(jobInfoStartingCell[1:])+5): project.equipment
}

# Defining thin border style
thinBorder = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
                                 
# Defining fill style for highlighting
highlightFill = PatternFill(start_color="FFFF00", end_color="00FF00", fill_type="solid")

# Format and highlight cells
for cell, value in cells.items():
    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal='center')
    ws[cell].border = thinBorder
    ws[cell].fill = highlightFill  # Apply highlight

ws.column_dimensions[jobInfoStartingCell[0]].width = 50;

nestedRows = [
    ["General",
        [
            "Item 1",
            ["subitem 1", "subitem 2", "subitem 3"]
        ],
        [
            "Item 2",
            ["subitem 1", "subitem 2", "subitem 3"]
        ]
    ],
    ["Conditions",
        [
            "Item 1",
            ["subitem 1", "subitem 2", "subitem 3"]
        ],
        [
            "Item 2",
            ["subitem 1", "subitem 2", "subitem 3"]
        ]
    ],
    ["Electrical",
        [
            "Item 1",
            ["subitem 1", "subitem 2", "subitem 3"]
        ],
        [
            "Item 2",
            ["subitem 1", "subitem 2", "subitem 3"]
        ]
    ]
]

count = 5

for category in nestedRows:
    print(category[0])  # Print the category name

    ws[dataCell]=category[0]; # populate data
    ws[dataCell].font=Font(bold=True) # bold data cell
    dataCell=increment_cell(dataCell) # increment data cell

    multiLevelNum = incrementMultiLevelNum(multiLevelNum,'main') # increment multilevel number 
    ws[multiLevelNumCell] = multiLevelNum; # populate multilevel number 

    ws[multiLevelNumCell].font=Font(bold=True) # bold multilevel num cell
    multiLevelNumCell=increment_cell(multiLevelNumCell) # increment multilevel cell 

    for item in category[1:]:  # Iterate through each item in the category
        print(item[0])  # Print the item name

        ws[dataCell]=item[0]; # populate data
        dataCell=increment_cell(dataCell) # increment data cell

        multiLevelNum = incrementMultiLevelNum(multiLevelNum,'sub')
        ws[multiLevelNumCell] = multiLevelNum; # populate multilevel number 
        ws[multiLevelNumCell].alignment = Alignment(horizontal='center')
        multiLevelNumCell=increment_cell(multiLevelNumCell) # increment multilevel cell 


        for subitem in item[1]:  # Iterate through each subitem in the item
            print(subitem)  # Print the subitem name

            ws[dataCell]=subitem; # populate data
            dataCell=increment_cell(dataCell)  # increment data cell

            multiLevelNum = incrementMultiLevelNum(multiLevelNum,'subsub') 
            ws[multiLevelNumCell] = multiLevelNum; # populate multilevel number 
            ws[multiLevelNumCell].alignment = Alignment(horizontal='right')
            multiLevelNumCell=increment_cell(multiLevelNumCell) # increment multilevel cell 


# Save the workbook to a file
wb.save("Quote.xlsx")
